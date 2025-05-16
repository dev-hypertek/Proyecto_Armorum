from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import os
import shutil
import json
from datetime import datetime
from typing import List, Optional
import tempfile
from openpyxl import Workbook
from dotenv import load_dotenv
import xml.etree.ElementTree as ET
import csv
import pandas as pd

# Cargar variables de entorno al inicio
load_dotenv()

# Firebase service - importación simplificada
from firebase_service import firebase_service

app = FastAPI(title="Armorum API", version="1.0.0")

# CORS - Allow all origins for development (update for production)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://armorum-financiero.web.app", "http://localhost:3000", "http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def detect_file_format_by_content(file_path: str, filename: str) -> str:
    """Detecta el formato del archivo por contenido, no solo por extensión"""
    print(f"[CONSOLE LOG] Detectando formato para archivo: {filename}")
    
    try:
        # Leer primeras líneas para detectar formato
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            first_lines = f.read(1024)  # Leer primeros 1KB
        
        # Detectar XML
        if first_lines.strip().startswith('<?xml') or '<' in first_lines:
            print(f"[CONSOLE LOG] Formato detectado por contenido: XML")
            return 'xml'
        
        # Detectar CSV por delimitadores
        if ',' in first_lines and ('\n' in first_lines or '\r' in first_lines):
            lines = first_lines.split('\n')[:3]
            comma_count = sum(line.count(',') for line in lines)
            if comma_count > 0:
                print(f"[CONSOLE LOG] Formato detectado por contenido: CSV")
                return 'csv_excel'
        
        # Detectar TXT estructurado (con múltiples espacios como separadores)
        if '    ' in first_lines or '\t' in first_lines:
            print(f"[CONSOLE LOG] Formato detectado por contenido: TXT estructurado")
            return 'txt_plano'
        
        # Fallback a extensión
        extension = filename.lower().split('.')[-1]
        print(f"[CONSOLE LOG] Usando extensión como fallback: {extension}")
        
        if extension == 'xml':
            return 'xml'
        elif extension in ['csv', 'xlsx', 'xls']:
            return 'csv_excel'
        elif extension == 'txt':
            return 'txt_plano'
        else:
            return 'unknown'
            
    except Exception as e:
        print(f"[CONSOLE LOG] Error detectando formato: {e}")
        return 'unknown'

def process_file_by_type(file_path: str, formato: str, filename: str):
    """Procesa archivos según su tipo (XML, CSV/Excel, TXT)"""
    print(f"[CONSOLE LOG] Procesando archivo: {filename} con formato: {formato}")
    
    # Si es auto_detect, usar detección inteligente por contenido
    if formato == 'auto_detect':
        formato = detect_file_format_by_content(file_path, filename)
        print(f"[CONSOLE LOG] Formato auto-detectado: {formato}")
    
    try:
        if formato == "xml":
            return process_xml_file(file_path, filename)
        elif formato in ["csv_excel", "plantilla51"]:
            extension = filename.lower().split('.')[-1]
            return process_csv_excel_file(file_path, extension, filename)
        elif formato == "txt_plano":
            return process_txt_file(file_path, filename)
        else:
            print(f"[CONSOLE LOG] ERROR: Formato no reconocido: {formato}")
            return {
                "success": False,
                "registros": 0,
                "errores": [f"Formato de archivo no reconocido: {formato}"]
            }
    except Exception as e:
        print(f"[CONSOLE LOG] ERROR en procesamiento: {str(e)}")
        return {
            "success": False,
            "registros": 0,
            "errores": [f"Error procesando archivo: {str(e)}"]
        }

def process_xml_file(file_path: str, filename: str):
    """Procesa archivos XML de facturas electrónicas"""
    print(f"[CONSOLE LOG] Iniciando procesamiento XML: {filename}")
    
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        print(f"[CONSOLE LOG] XML parseado exitosamente. Root element: {root.tag}")
        
        # Buscar facturas de manera más precisa
        facturas_encontradas = 0
        errores = []
        
        # Patrón 1: Buscar elementos con "Factura" en el nombre (caso común)
        facturas_found = root.findall(".//FacturaElectronica")
        if facturas_found:
            facturas_encontradas = len(facturas_found)
            print(f"[CONSOLE LOG] Encontradas {facturas_encontradas} FacturaElectronica")
        
        # Patrón 2: Si no encuentra, buscar por otros patrones comunes
        if facturas_encontradas == 0:
            facturas_found = root.findall(".//Factura")
            if facturas_found:
                facturas_encontradas = len(facturas_found)
                print(f"[CONSOLE LOG] Encontradas {facturas_encontradas} Factura")
        
        # Patrón 3: Si no encuentra, buscar por Invoice (formato internacional)
        if facturas_encontradas == 0:
            facturas_found = root.findall(".//Invoice")
            if facturas_found:
                facturas_encontradas = len(facturas_found)
                print(f"[CONSOLE LOG] Encontradas {facturas_encontradas} Invoice")
        
        # Si aún no encuentra facturas, verificar si el root es una factura
        if facturas_encontradas == 0:
            if "factura" in root.tag.lower() or "invoice" in root.tag.lower():
                facturas_encontradas = 1
                print(f"[CONSOLE LOG] Root element es una factura: {root.tag}")
            else:
                # Como último recurso, contar elementos hijos directos que podrían ser facturas
                children = list(root)
                if children:
                    facturas_encontradas = len(children)
                    print(f"[CONSOLE LOG] Contando elementos hijos como facturas: {facturas_encontradas}")
                else:
                    facturas_encontradas = 1  # Al menos hay un XML válido
                    print(f"[CONSOLE LOG] XML válido pero sin estructura de facturas clara")
        
        # Validar datos mínimos en las facturas encontradas
        if facturas_encontradas > 0:
            # Verificar que al menos una factura tenga datos
            sample_factura = facturas_found[0] if facturas_found else root
            
            # Buscar elementos comunes en facturas
            numero_factura = sample_factura.find(".//NumeroFactura")
            fecha = sample_factura.find(".//FechaEmision") or sample_factura.find(".//Fecha")
            total = sample_factura.find(".//Total") or sample_factura.find(".//TotalFactura")
            
            if not any([numero_factura, fecha, total]):
                errores.append("XML válido pero sin estructura de factura reconocible")
                print(f"[CONSOLE LOG] WARNING: No se encontraron datos de factura estándar")
        
        print(f"[CONSOLE LOG] Procesamiento XML completado: {facturas_encontradas} facturas")
        return {
            "success": True,
            "registros": facturas_encontradas,
            "errores": errores,
            "tipo": "XML",
            "detalles": {
                "rootElement": root.tag,
                "namespaces": list(root.nsmap.keys()) if hasattr(root, 'nsmap') else []
            }
        }
        
    except ET.ParseError as e:
        error_msg = f"Error parsing XML: {str(e)}"
        print(f"[CONSOLE LOG] ERROR: {error_msg}")
        return {
            "success": False,
            "registros": 0,
            "errores": [error_msg]
        }
    except Exception as e:
        error_msg = f"Error inesperado procesando XML: {str(e)}"
        print(f"[CONSOLE LOG] ERROR: {error_msg}")
        return {
            "success": False,
            "registros": 0,
            "errores": [error_msg]
        }

def process_csv_excel_file(file_path: str, extension: str, filename: str):
    """Procesa archivos CSV y Excel"""
    print(f"[CONSOLE LOG] Iniciando procesamiento {extension.upper()}: {filename}")
    
    try:
        # Intentar diferentes encodings para CSV
        if extension == "csv":
            # Primero intentar con UTF-8
            try:
                df = pd.read_csv(file_path, encoding='utf-8')
                print(f"[CONSOLE LOG] CSV leído con encoding UTF-8")
            except UnicodeDecodeError:
                # Si falla, intentar con latin-1
                df = pd.read_csv(file_path, encoding='latin-1')
                print(f"[CONSOLE LOG] CSV leído con encoding latin-1")
        else:
            # Leer Excel
            df = pd.read_excel(file_path)
            print(f"[CONSOLE LOG] Excel leído exitosamente")
        
        # Verificar si tiene datos
        registros = len(df)
        errores = []
        print(f"[CONSOLE LOG] Total de registros encontrados: {registros}")
        print(f"[CONSOLE LOG] Columnas encontradas: {list(df.columns)}")
        
        # Validaciones básicas
        if registros == 0:
            errores.append("El archivo no contiene datos")
            print(f"[CONSOLE LOG] ERROR: Archivo vacío")
        
        # Verificar si tiene columnas
        if len(df.columns) == 0:
            errores.append("El archivo no tiene columnas definidas")
            print(f"[CONSOLE LOG] ERROR: Sin columnas definidas")
        
        # Validaciones específicas para archivos de factura
        column_names = [col.strip().upper() for col in df.columns]
        expected_columns = [
            'NOMBRE USUARIO', 'NIT USUARIO', 'FACT NRO', 'FECHA', 
            'PRODUCTO', 'CANTIDAD', 'VALOR UNITARIO', 'TOTAL'
        ]
        
        # Verificar si parece un archivo de facturas
        is_invoice_file = False
        for exp_col in expected_columns[:4]:  # Verificar al menos las primeras 4 columnas
            if any(exp_col in col for col in column_names):
                is_invoice_file = True
                break
        
        if not is_invoice_file and registros > 0:
            print(f"[CONSOLE LOG] WARNING: Archivo no parece tener estructura de facturas estándar")
        
        # Detectar filas vacías
        empty_rows = df.isnull().all(axis=1).sum()
        if empty_rows > 0:
            print(f"[CONSOLE LOG] WARNING: {empty_rows} filas vacías encontradas")
        
        # Información adicional
        sample_data = df.head(3).to_dict('records') if registros > 0 else []
        
        print(f"[CONSOLE LOG] Procesamiento {extension.upper()} completado exitosamente")
        return {
            "success": len(errores) == 0,
            "registros": registros,
            "errores": errores,
            "tipo": f"CSV/Excel ({extension.upper()})",
            "columnas": list(df.columns)[:15],  # Primeras 15 columnas
            "detalles": {
                "emptyRows": empty_rows,
                "isInvoiceFormat": is_invoice_file,
                "sampleData": sample_data
            }
        }
        
    except pd.errors.EmptyDataError:
        error_msg = "El archivo está vacío o no contiene datos válidos"
        print(f"[CONSOLE LOG] ERROR: {error_msg}")
        return {
            "success": False,
            "registros": 0,
            "errores": [error_msg]
        }
    except pd.errors.ParserError as e:
        error_msg = f"Error parseando {extension.upper()}: {str(e)}"
        print(f"[CONSOLE LOG] ERROR: {error_msg}")
        return {
            "success": False,
            "registros": 0,
            "errores": [error_msg]
        }
    except Exception as e:
        error_msg = f"Error procesando {extension.upper()}: {str(e)}"
        print(f"[CONSOLE LOG] ERROR: {error_msg}")
        return {
            "success": False,
            "registros": 0,
            "errores": [error_msg]
        }

def process_txt_file(file_path: str, filename: str):
    """Procesa archivos de texto plano con estructura de facturación"""
    print(f"[CONSOLE LOG] Iniciando procesamiento TXT: {filename}")
    
    # Estructura estándar esperada (15 columnas)
    EXPECTED_COLUMNS = [
        'NOMBRE VENDEDOR', 'NIT VENDEDOR', 'NOMBRE COMPRADOR', 'NIT COMPRADOR',
        'CIUDAD DE ENTREGA DEL PRODUCTO', 'FACT NRO', 'PREFIJO', 'FECHA',
        'FORMA DE PAGO(#DIAS)', 'PRODUCTO', 'CANTIDAD', 'UNIDAD',
        'VALOR UNITARIO', 'TOTAL', '% IVA PRODUCTO'
    ]
    
    try:
        # Intentar diferentes encodings
        content = None
        encoding_used = None
        
        for encoding in ['utf-8', 'latin-1', 'cp1252']:
            try:
                with open(file_path, 'r', encoding=encoding) as file:
                    content = file.read()
                    encoding_used = encoding
                    print(f"[CONSOLE LOG] TXT leído con encoding {encoding}")
                    break
            except UnicodeDecodeError:
                continue
        
        if content is None:
            raise Exception("No se pudo leer el archivo con ningún encoding")
        
        lines = content.split('\n')
        
        # Filtrar líneas no vacías y quitar espacios extra
        non_empty_lines = [line.strip() for line in lines if line.strip()]
        total_lines = len(non_empty_lines)
        print(f"[CONSOLE LOG] Total de líneas no vacías: {total_lines}")
        
        errores = []
        if total_lines == 0:
            errores.append("El archivo de texto está vacío")
            print(f"[CONSOLE LOG] ERROR: Archivo vacío")
            return {
                "success": False,
                "registros": 0,
                "errores": errores
            }
        
        # Analizar estructura del archivo
        estructura_detectada = "Texto libre"
        registros_datos = 0
        separador_detectado = None
        
        # Buscar línea de cabecera con las columnas estándar
        header_line_idx = -1
        for i, line in enumerate(non_empty_lines):
            # Verificar si contiene columnas estándar de facturación
            line_upper = line.upper()
            if 'NOMBRE VENDEDOR' in line_upper and 'NIT VENDEDOR' in line_upper:
                header_line_idx = i
                print(f"[CONSOLE LOG] Header encontrado en línea {i}")
                break
            elif any(keyword in line_upper for keyword in ['FACT NRO', 'FECHA', 'PRODUCTO']):
                header_line_idx = i
                print(f"[CONSOLE LOG] Posible header alternativo en línea {i}")
                break
        
        # Detectar separador basado en la línea de header o primera línea de datos
        sample_line = None
        if header_line_idx >= 0 and header_line_idx + 1 < len(non_empty_lines):
            sample_line = non_empty_lines[header_line_idx + 1]  # Primera línea de datos
        else:
            # Buscar primera línea que parezca datos
            for line in non_empty_lines:
                if not line.startswith('=') and not line.startswith('-') and \
                   not any(keyword in line.upper() for keyword in ['FACTURAS', 'ARMORUM', 'FORMATO']):
                    sample_line = line
                    break
        
        if sample_line:
            print(f"[CONSOLE LOG] Línea de muestra: {sample_line[:100]}...")
            
            # Detectar separadores
            if '\t' in sample_line:
                estructura_detectada = "Delimitado por tabs"
                separador_detectado = '\t'
                print(f"[CONSOLE LOG] Detectado delimitador TAB")
            elif '|' in sample_line:
                estructura_detectada = "Delimitado por |"
                separador_detectado = '|'
                print(f"[CONSOLE LOG] Detectado delimitador |")
            elif sample_line.count(',') >= 5:  # Al menos 5 comas para 15 columnas
                estructura_detectada = "Posiblemente CSV"
                separador_detectado = ','
                print(f"[CONSOLE LOG] Detectado delimitador ,")
            elif sample_line.count('   ') >= 3:  # Múltiples espacios para separar campos
                estructura_detectada = "Delimitado por espacios múltiples"
                separador_detectado = 'espacios'
                print(f"[CONSOLE LOG] Detectado delimitador por espacios múltiples")
        
        # Contar registros de datos
        start_idx = max(header_line_idx + 1, 0) if header_line_idx >= 0 else 0
        registros_datos = 0
        
        for i in range(start_idx, len(non_empty_lines)):
            line = non_empty_lines[i]
            
            # Saltar líneas que claramente no son datos
            if line.startswith('=') or line.startswith('-') or \
               any(keyword in line.upper() for keyword in ['FACTURAS', 'ARMORUM', 'FORMATO', 'NOVEDADES']):
                continue
            
            # Validar que la línea tenga la estructura esperada
            if separador_detectado and separador_detectado != 'espacios':
                campos = line.split(separador_detectado)
                if len(campos) >= 10:  # Al menos 10 campos de los 15 esperados
                    registros_datos += 1
            elif separador_detectado == 'espacios':
                # Para espacios múltiples, usar expresión regular para separar
                import re
                campos = re.split(r'\s{2,}', line)
                if len(campos) >= 10:
                    registros_datos += 1
            else:
                # Sin separador claro, asumir formato libre
                registros_datos += 1
        
        print(f"[CONSOLE LOG] Registros de datos detectados: {registros_datos}")
        print(f"[CONSOLE LOG] Estructura detectada: {estructura_detectada}")
        
        # Validar estructura estándar si se detectó header
        if header_line_idx >= 0:
            header_line = non_empty_lines[header_line_idx]
            matched_columns = 0
            for expected_col in EXPECTED_COLUMNS[:10]:  # Verificar al menos 10 de las 15 columnas
                if expected_col.upper() in header_line.upper():
                    matched_columns += 1
            
            if matched_columns < 5:
                errores.append(f"Estructura no coincide con formato estándar de facturación. Solo {matched_columns} columnas reconocidas")
                print(f"[CONSOLE LOG] WARNING: Estructura no estándar")
        
        # Extraer muestra de datos para análisis
        sample_records = []
        if registros_datos > 0 and separador_detectado:
            for i in range(min(3, registros_datos)):
                line_idx = start_idx + i
                if line_idx < len(non_empty_lines):
                    line = non_empty_lines[line_idx]
                    if separador_detectado == 'espacios':
                        import re
                        campos = re.split(r'\s{2,}', line)
                    else:
                        campos = line.split(separador_detectado)
                    sample_records.append(campos[:15])  # Máximo 15 campos
        
        print(f"[CONSOLE LOG] Procesamiento TXT completado")
        return {
            "success": len(errores) == 0,
            "registros": registros_datos,
            "errores": errores,
            "tipo": "TXT",
            "estructura": estructura_detectada,
            "detalles": {
                "separador": separador_detectado,
                "headerEncontrado": header_line_idx >= 0,
                "totalLineas": total_lines,
                "estructuraEstandar": header_line_idx >= 0,
                "sampleRecords": sample_records,
                "encoding": encoding_used
            }
        }
        
    except Exception as e:
        error_msg = f"Error procesando TXT: {str(e)}"
        print(f"[CONSOLE LOG] ERROR: {error_msg}")
        return {
            "success": False,
            "registros": 0,
            "errores": [error_msg]
        }

@app.get("/")
async def root():
    return {"message": "Armorum API funcionando en Firebase"}

@app.post("/api/facturas/cargar")
async def cargar_archivo(
    archivo: UploadFile = File(...),
    clienteId: str = Form(...),
    formatoArchivo: str = Form(...)
):
    print(f"[CONSOLE LOG] ==== INICIO CARGA ARCHIVO ====")
    print(f"[CONSOLE LOG] Archivo: {archivo.filename}")
    print(f"[CONSOLE LOG] Cliente ID: {clienteId}")
    print(f"[CONSOLE LOG] Formato solicitado: {formatoArchivo}")
    print(f"[CONSOLE LOG] Tamaño archivo: {archivo.size} bytes")
    
    # Validaciones básicas
    if archivo.size > 50 * 1024 * 1024:  # 50MB
        print(f"[CONSOLE LOG] ERROR: Archivo muy grande ({archivo.size} bytes)")
        raise HTTPException(status_code=413, detail="Archivo muy grande")
    
    # Validar extensión
    allowed_extensions = ['.xml', '.csv', '.xlsx', '.xls', '.txt']
    file_extension = os.path.splitext(archivo.filename)[1].lower()
    if file_extension not in allowed_extensions:
        print(f"[CONSOLE LOG] WARNING: Extensión no estándar: {file_extension}")
    
    # Guardar archivo temporalmente
    file_path = f"uploads/{archivo.filename}"
    os.makedirs("uploads", exist_ok=True)
    print(f"[CONSOLE LOG] Guardando archivo en: {file_path}")
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(archivo.file, buffer)
    print(f"[CONSOLE LOG] Archivo guardado exitosamente")
    
    # Mapeo de clientes
    cliente_nombres = {"1": "Comiagro", "2": "Olímpica", "3": "Cliente Regional"}
    cliente_seleccionado = cliente_nombres.get(clienteId, "Cliente Desconocido")
    print(f"[CONSOLE LOG] Cliente seleccionado: {cliente_seleccionado}")
    
    # Detección automática de formato si es 'auto_detect'
    formato_final = formatoArchivo
    if formatoArchivo == 'auto_detect':
        print(f"[CONSOLE LOG] Iniciando detección automática de formato")
        formato_final = detect_file_format_by_content(file_path, archivo.filename)
        print(f"[CONSOLE LOG] Formato detectado automáticamente: {formato_final}")
    
    # Crear lote en Firestore
    print(f"[CONSOLE LOG] Creando lote en Firestore")
    lote_data = {
        "nombreArchivo": archivo.filename,
        "cliente": cliente_seleccionado,
        "formato": formato_final,
        "estado": "Procesando",
        "registrosTotales": 0,
        "errores": 0,
        "filePath": file_path
    }
    
    lote_id = firebase_service.create_lote(lote_data)
    print(f"[CONSOLE LOG] Lote creado con ID: {lote_id}")
    firebase_service.add_log(lote_id, f"Archivo recibido. Tipo: {formato_final}. Cliente: {cliente_seleccionado}")
    
    # Procesar archivo según tipo
    try:
        print(f"[CONSOLE LOG] Iniciando procesamiento del archivo")
        resultado_procesamiento = process_file_by_type(file_path, formato_final, archivo.filename)
        print(f"[CONSOLE LOG] Resultado procesamiento: {resultado_procesamiento}")
        
        if resultado_procesamiento["success"]:
            registros_totales = resultado_procesamiento["registros"]
            print(f"[CONSOLE LOG] Procesamiento exitoso. Registros: {registros_totales}")
            
            # Simular errores realistas basados en las reglas de negocio
            errores_count = 0
            errores_simulados = []
            
            if registros_totales > 0:
                # Simulación más inteligente de errores
                import random
                random.seed(42)  # Para reproducibilidad en testing
                
                # Calcular número de errores (entre 3% y 7% para ser realista)
                error_rate = random.uniform(0.03, 0.07)
                num_errores = max(1, int(registros_totales * error_rate))
                print(f"[CONSOLE LOG] Simulando {num_errores} errores ({error_rate:.1%} de {registros_totales} registros)")
                
                # Generar filas con errores (distribuidas aleatoriamente)
                filas_con_error = random.sample(range(1, registros_totales + 1), min(num_errores, registros_totales))
                
                for idx, fila in enumerate(filas_con_error):
                    # Tipos de errores basados en las reglas de negocio
                    error_types = [
                        {
                            "campo": "PRODUCTO",
                            "mensaje": f"Fila {fila}: Producto '{generate_product_name()}' no encontrado en catálogo BMC",
                            "severidad": "WARNING",
                            "create_exception": False
                        },
                        {
                            "campo": "NIT_COMPRADOR",
                            "mensaje": f"Fila {fila}: NIT comprador no encontrado en DIAN",
                            "severidad": "ERROR",
                            "create_exception": True
                        },
                        {
                            "campo": "CANTIDAD",
                            "mensaje": f"Fila {fila}: Cantidad no especificada en kilos/litros. Requiere conversión manual",
                            "severidad": "WARNING",
                            "create_exception": False
                        },
                        {
                            "campo": "% IVA PRODUCTO",
                            "mensaje": f"Fila {fila}: IVA debe ser número entero (0, 19, 5). Valor actual inválido",
                            "severidad": "ERROR",
                            "create_exception": False
                        },
                        {
                            "campo": "FECHA",
                            "mensaje": f"Fila {fila}: Fecha debe estar en formato DD/MM/AAAA",
                            "severidad": "WARNING",
                            "create_exception": False
                        },
                        {
                            "campo": "NIT_VENDEDOR",
                            "mensaje": f"Fila {fila}: NIT vendedor no encontrado en DIAN",
                            "severidad": "ERROR",
                            "create_exception": True
                        }
                    ]
                    
                    # Seleccionar tipo de error de manera balanceada
                    error_tipo = error_types[idx % len(error_types)]
                    
                    print(f"[CONSOLE LOG] Error simulado - Fila {fila}: {error_tipo['campo']} - {error_tipo['severidad']}")
                    
                    # Agregar error
                    firebase_service.add_error(
                        lote_id,
                        fila,
                        error_tipo["campo"],
                        error_tipo["mensaje"],
                        error_tipo["severidad"]
                    )
                    
                    errores_count += 1
                    
                    # Crear excepción DIAN si es necesario
                    if error_tipo["create_exception"]:
                        nit_type = "vendedor" if "VENDEDOR" in error_tipo["campo"] else "comprador"
                        excepcion_data = {
                            "loteId": lote_id,
                            "filaOrigen": fila,
                            "documento": f"{random.randint(800000000, 999999999)}-{random.randint(1,9)}",
                            "nombreReportado": f"Empresa {fila} SAS",
                            "estadoValidacion": random.choice(["No_Encontrado", "Inconsistente"]),
                            "estadoGestion": "Pendiente",
                            "notas": f"Error en validación de NIT {nit_type}",
                            "tipoTercero": nit_type
                        }
                        firebase_service.create_excepcion_dian(excepcion_data)
                        print(f"[CONSOLE LOG] Excepción DIAN creada para fila {fila}")
            
            # Determinar estado final
            if errores_count == 0:
                estado = "Completado"
            elif errores_count <= registros_totales * 0.1:  # Menos del 10% son advertencias
                estado = "Completado con Advertencias"
            else:
                estado = "Error"
            
            print(f"[CONSOLE LOG] Estado final: {estado} ({errores_count} errores)")
            
        else:
            # Error en el procesamiento del archivo
            print(f"[CONSOLE LOG] ERROR en procesamiento: {resultado_procesamiento['errores']}")
            registros_totales = 0
            errores_count = len(resultado_procesamiento["errores"])
            estado = "Error"
            
            for error in resultado_procesamiento["errores"]:
                firebase_service.add_error(lote_id, 0, "ARCHIVO", error, "ERROR")
        
        # Actualizar lote con resultados del procesamiento
        print(f"[CONSOLE LOG] Actualizando lote en Firestore")
        firebase_service.update_lote(lote_id, {
            "registrosTotales": registros_totales,
            "errores": errores_count,
            "estado": estado,
            "tipoArchivoDetectado": resultado_procesamiento.get("tipo", formato_final)
        })
        
        firebase_service.add_log(lote_id, f"Procesamiento completado. Estado: {estado}. Registros: {registros_totales}. Errores: {errores_count}")
        
    except Exception as e:
        print(f"[CONSOLE LOG] EXCEPCIÓN en procesamiento: {str(e)}")
        firebase_service.update_lote(lote_id, {"estado": "Error", "errores": 1})
        firebase_service.add_log(lote_id, f"Error en procesamiento: {str(e)}", "ERROR")
        estado = "Error"
        registros_totales = 0
        errores_count = 1
    
    # Limpiar archivo temporal
    try:
        os.remove(file_path)
        print(f"[CONSOLE LOG] Archivo temporal eliminado")
    except Exception as e:
        print(f"[CONSOLE LOG] WARNING: No se pudo eliminar archivo temporal: {e}")
    
    print(f"[CONSOLE LOG] ==== FIN CARGA ARCHIVO ====")
    return {
        "success": True,
        "message": "Archivo recibido y procesado",
        "data": {
            "loteId": lote_id,
            "nombreArchivo": archivo.filename,
            "estado": estado,
            "fechaCarga": datetime.utcnow().isoformat(),
            "registrosEstimados": registros_totales,
            "erroresEncontrados": errores_count,
            "formatoDetectado": resultado_procesamiento.get("tipo", formato_final) if 'resultado_procesamiento' in locals() else formato_final
        }
    }

# Función auxiliar para generar nombres de productos realistas
def generate_product_name():
    productos = [
        "LECHUGA ROMANA", "PAPA CRIOLLA", "AREPA BLANCA", "QUESO MOZARELLA",
        "TOMATE CHONTO", "CEBOLLA CABEZONA", "ZANAHORIA", "CILANTRO",
        "BANANO BOCADILLO", "MANGO TOMMY", "AGUACATE HASS", "LIMÓN TAHITÍ"
    ]
    import random
    return random.choice(productos)

# Resto de endpoints sin cambios...
@app.get("/api/facturas/lotes")
async def obtener_lotes(
    page: int = 1,
    limit: int = 20,
    estado: Optional[str] = None
):
    # Get lotes from Firestore
    lotes = firebase_service.get_lotes(limit=limit * page)  # Simple pagination
    
    # Filter by estado if provided
    if estado:
        lotes = [l for l in lotes if l.get("estado") == estado]
    
    # Simple pagination (can be improved with cursor-based pagination)
    start = (page - 1) * limit
    end = start + limit
    lotes_pagina = lotes[start:end]
    
    # Convert Firestore timestamps to ISO format
    for lote in lotes_pagina:
        if 'fechaCarga' in lote:
            lote['fechaCarga'] = lote['fechaCarga'].isoformat()
        if 'fechaUltimaActualizacion' in lote:
            lote['fechaUltimaActualizacion'] = lote['fechaUltimaActualizacion'].isoformat()
    
    return {
        "success": True,
        "data": {
            "lotes": lotes_pagina,
            "paginacion": {
                "paginaActual": page,
                "totalPaginas": (len(lotes) + limit - 1) // limit,
                "totalRegistros": len(lotes),
                "registrosPorPagina": limit
            }
        }
    }

@app.get("/api/facturas/lotes/{lote_id}")
async def obtener_detalle_lote(lote_id: str):
    try:
        lote = firebase_service.get_lote_by_id(lote_id)
        if not lote:
            raise HTTPException(status_code=404, detail="Lote no encontrado")
    except Exception as e:
        print(f"Error getting lote: {e}")
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")
    
    # Get logs and errors
    logs = firebase_service.get_logs_by_lote(lote_id)
    errores = firebase_service.get_errors_by_lote(lote_id)
    
    # Convert timestamps
    for log in logs:
        if 'timestamp' in log:
            log['timestamp'] = log['timestamp'].isoformat()
    
    for error in errores:
        if 'fechaDeteccion' in error:
            error['fechaDeteccion'] = error['fechaDeteccion'].isoformat()
    
    # Convert lote timestamps
    if 'fechaCarga' in lote:
        lote['fechaCarga'] = lote['fechaCarga'].isoformat()
    if 'fechaUltimaActualizacion' in lote:
        lote['fechaUltimaActualizacion'] = lote['fechaUltimaActualizacion'].isoformat()
    
    # Agrupar errores por tipo para mejor visualización
    errores_por_tipo = {}
    for error in errores:
        tipo = error.get('campo', 'OTROS')
        if tipo not in errores_por_tipo:
            errores_por_tipo[tipo] = []
        errores_por_tipo[tipo].append(error)
    
    return {
        "success": True,
        "data": {
            "lote": lote,
            "logs": logs,
            "errores": errores,
            "erroresPorTipo": errores_por_tipo,
            "puedeDescargar": lote.get("estado") in ["Completado", "Completado con Advertencias"]
        }
    }

@app.get("/api/facturas/lotes/{lote_id}/descargar")
async def descargar_plantilla(lote_id: str):
    lote = firebase_service.get_lote_by_id(lote_id)
    if not lote or lote.get("estado") not in ["Completado", "Completado con Advertencias"]:
        raise HTTPException(status_code=404, detail="Plantilla no disponible")
    
    # Generate Excel file based on PLANTILLA51_SIMONA structure
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Comiagro"
    
    # Headers based on PLANTILLA51_SIMONA
    headers = [
        "NOMBRE USUARIO (CLIENTE DEL CLIENTE)",
        "NIT USUARIO (CLIENTE DEL CLIENTE)", 
        "CIUDAD DE ENTREGA DEL PRODUCTO",
        "FACT NRO",
        "FECHA",
        "FORMA DE PAGO",
        "PRODUCTO",
        "PRESENTACION",
        "CANTIDAD",
        "VALOR UNITARIO",
        "TOTAL",
        "% IVA PRODUCTO"
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add sample data based on the lote
    registros = lote.get('registrosTotales', 1)
    for row in range(2, min(registros + 2, 10)):  # Máximo 8 filas de ejemplo
        ws.cell(row=row, column=1, value=f"Cliente {row-1} SAS")
        ws.cell(row=row, column=2, value=f"9012345{row-1:02d}")
        ws.cell(row=row, column=3, value="Bogotá")
        ws.cell(row=row, column=4, value=f"F-{row-1:04d}")
        ws.cell(row=row, column=5, value=datetime.now().strftime("%Y-%m-%d"))
        ws.cell(row=row, column=6, value="CREDITO")
        ws.cell(row=row, column=7, value=f"PRODUCTO_{row-1}")
        ws.cell(row=row, column=8, value="KG")
        ws.cell(row=row, column=9, value=10)
        ws.cell(row=row, column=10, value=25000)
        ws.cell(row=row, column=11, value=250000)
        ws.cell(row=row, column=12, value=19)
    
    # Save to temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        temp_path = tmp.name
    
    return FileResponse(
        temp_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"plantilla_comiagro_lote_{lote_id}.xlsx"
    )

@app.get("/api/terceros/excepciones")
async def obtener_excepciones(
    estado: Optional[str] = None,
    lote_id: Optional[str] = None
):
    filters = {}
    if estado:
        filters['estado'] = estado
    if lote_id:
        filters['lote_id'] = lote_id
    
    excepciones = firebase_service.get_excepciones_dian(filters)
    
    # Convert timestamps
    for exc in excepciones:
        if 'fechaDeteccion' in exc:
            exc['fechaDeteccion'] = exc['fechaDeteccion'].isoformat()
        if 'fechaUltimaActualizacion' in exc:
            exc['fechaUltimaActualizacion'] = exc['fechaUltimaActualizacion'].isoformat()
    
    # Calculate statistics
    total = len(excepciones)
    pendientes = len([e for e in excepciones if e.get('estadoGestion') == 'Pendiente'])
    resueltas = total - pendientes
    
    return {
        "success": True,
        "data": {
            "excepciones": excepciones,
            "estadisticas": {
                "total": total,
                "pendientes": pendientes,
                "resueltas": resueltas
            }
        }
    }

@app.post("/api/terceros/excepciones/{excepcion_id}/{action}")
async def actualizar_estado_tercero(
    excepcion_id: str,
    action: str,
    payload: Optional[dict] = None
):
    # Get excepcion from Firestore
    excepciones = firebase_service.get_excepciones_dian()
    excepcion = next((e for e in excepciones if e.get('id') == excepcion_id), None)
    
    if not excepcion:
        raise HTTPException(status_code=404, detail="Excepción no encontrada")
    
    estado_anterior = excepcion.get('estadoValidacion')
    
    # Mapear acciones a estados
    accion_map = {
        "corregir": "Corregida",
        "crear": "En_Creacion_Manual",
        "ignorar": "Ignorada",
        "reintentar": "Reintentando"
    }
    
    if action not in accion_map:
        raise HTTPException(status_code=400, detail="Acción no válida")
    
    updates = {
        "estadoGestion": accion_map[action]
    }
    
    if payload:
        if 'notas' in payload:
            updates['notas'] = payload['notas']
        if 'datosCorreccion' in payload:
            updates['datosCorreccion'] = payload['datosCorreccion']
    
    # Update in Firestore
    success = firebase_service.update_excepcion_dian(excepcion_id, updates)
    
    if not success:
        raise HTTPException(status_code=500, detail="Error al actualizar excepción")
    
    return {
        "success": True,
        "message": f"Excepción actualizada: {action}",
        "data": {
            "excepcionId": excepcion_id,
            "accionAplicada": action,
            "estadoAnterior": estado_anterior,
            "estadoNuevo": accion_map[action],
            "fechaActualizacion": datetime.utcnow().isoformat()
        }
    }

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8080))
    uvicorn.run(app, host="0.0.0.0", port=port)