from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import os
import shutil
import json
from datetime import datetime
from typing import List, Optional
import pandas as pd
import tempfile
from openpyxl import Workbook

# Firebase service
from app.services.firebase_service import firebase_service

app = FastAPI(title="Armorum API", version="1.0.0")

# CORS - Allow all origins for development (update for production)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Update with specific domains in production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
async def root():
    return {"message": "Armorum API funcionando en Firebase"}

@app.post("/api/facturas/cargar")
async def cargar_archivo(
    archivo: UploadFile = File(...),
    clienteId: str = Form(...),
    formatoArchivo: str = Form(...)
):
    # Validaciones básicas
    if archivo.size > 50 * 1024 * 1024:  # 50MB
        raise HTTPException(status_code=413, detail="Archivo muy grande")
    
    # Guardar archivo temporalmente
    file_path = f"uploads/{archivo.filename}"
    os.makedirs("uploads", exist_ok=True)
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(archivo.file, buffer)
    
    # Mapeo de clientes
    cliente_nombres = {"1": "Comiagro", "2": "Cliente B", "3": "Cliente C"}
    
    # Crear lote en Firestore
    lote_data = {
        "nombreArchivo": archivo.filename,
        "cliente": cliente_nombres.get(clienteId, "Cliente Desconocido"),
        "formato": formatoArchivo,
        "estado": "Procesando",
        "registrosTotales": 0,
        "errores": 0,
        "filePath": file_path
    }
    
    lote_id = firebase_service.create_lote(lote_data)
    firebase_service.add_log(lote_id, "Archivo recibido y lote creado")
    
    # Procesar archivo de forma simple
    try:
        if archivo.filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file_path)
        elif archivo.filename.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.DataFrame()  # Para archivos .txt crear DataFrame vacío
        
        registros_totales = len(df) if not df.empty else 50
        estado = "Completado"
        errores = 0
        
        # Simular errores según formato
        if formatoArchivo == "plantilla51":
            errores = min(5, registros_totales // 10) if not df.empty else 3
            estado = "Error" if errores > 0 else "Completado"
            
            # Crear excepciones DIAN para demostración
            for i in range(min(3, errores)):
                excepcion_data = {
                    "loteId": lote_id,
                    "filaOrigen": 10 + i * 15,
                    "documento": f"90123456{i}",
                    "nombreReportado": f"Empresa {i+1} SAS",
                    "estadoValidacion": "No_Encontrado" if i % 2 == 0 else "Inconsistente",
                    "estadoGestion": "Pendiente",
                    "notas": None
                }
                firebase_service.create_excepcion_dian(excepcion_data)
                
                # Agregar error al lote
                firebase_service.add_error(
                    lote_id, 
                    10 + i * 15,
                    "NIT_COMPRADOR" if i % 2 == 0 else "NOMBRE_PRODUCTO",
                    "Tercero no encontrado en DIAN" if i % 2 == 0 else "Producto sin código BMC",
                    "ERROR"
                )
        
        # Actualizar lote con resultados del procesamiento
        firebase_service.update_lote(lote_id, {
            "registrosTotales": registros_totales,
            "errores": errores,
            "estado": estado
        })
        
        firebase_service.add_log(lote_id, f"Procesamiento completado. Estado: {estado}")
        
    except Exception as e:
        firebase_service.update_lote(lote_id, {"estado": "Error", "errores": 1})
        firebase_service.add_log(lote_id, f"Error en procesamiento: {str(e)}", "ERROR")
        estado = "Error"
        registros_totales = 0
    
    # Limpiar archivo temporal
    try:
        os.remove(file_path)
    except:
        pass
    
    return {
        "success": True,
        "message": "Archivo recibido y procesado",
        "data": {
            "loteId": lote_id,
            "nombreArchivo": archivo.filename,
            "estado": estado,
            "fechaCarga": datetime.utcnow().isoformat(),
            "registrosEstimados": registros_totales
        }
    }

@app.get("/api/facturas/lotes")
async def obtener_lotes(
    page: int = 1,
    limit: int = 20,
    estado: Optional[str] = None
):
    # Get lotes from Firestore
    lotes = firebase_service.get_lotes(limit=limit * page)  # Simple pagination
    
    # Filter by estado if provided
    if estado:
        lotes = [l for l in lotes if l.get("estado") == estado]
    
    # Simple pagination (can be improved with cursor-based pagination)
    start = (page - 1) * limit
    end = start + limit
    lotes_pagina = lotes[start:end]
    
    # Convert Firestore timestamps to ISO format
    for lote in lotes_pagina:
        if 'fechaCarga' in lote:
            lote['fechaCarga'] = lote['fechaCarga'].isoformat()
        if 'fechaUltimaActualizacion' in lote:
            lote['fechaUltimaActualizacion'] = lote['fechaUltimaActualizacion'].isoformat()
    
    return {
        "success": True,
        "data": {
            "lotes": lotes_pagina,
            "paginacion": {
                "paginaActual": page,
                "totalPaginas": (len(lotes) + limit - 1) // limit,
                "totalRegistros": len(lotes),
                "registrosPorPagina": limit
            }
        }
    }

@app.get("/api/facturas/lotes/{lote_id}")
async def obtener_detalle_lote(lote_id: str):
    lote = firebase_service.get_lote_by_id(lote_id)
    if not lote:
        raise HTTPException(status_code=404, detail="Lote no encontrado")
    
    # Get logs and errors
    logs = firebase_service.get_logs_by_lote(lote_id)
    errores = firebase_service.get_errors_by_lote(lote_id)
    
    # Convert timestamps
    for log in logs:
        if 'timestamp' in log:
            log['timestamp'] = log['timestamp'].isoformat()
    
    for error in errores:
        if 'fechaDeteccion' in error:
            error['fechaDeteccion'] = error['fechaDeteccion'].isoformat()
    
    # Convert lote timestamps
    if 'fechaCarga' in lote:
        lote['fechaCarga'] = lote['fechaCarga'].isoformat()
    if 'fechaUltimaActualizacion' in lote:
        lote['fechaUltimaActualizacion'] = lote['fechaUltimaActualizacion'].isoformat()
    
    return {
        "success": True,
        "data": {
            "lote": lote,
            "logs": logs,
            "errores": errores,
            "puedeDescargar": lote.get("estado") == "Completado"
        }
    }

@app.get("/api/facturas/lotes/{lote_id}/descargar")
async def descargar_plantilla(lote_id: str):
    lote = firebase_service.get_lote_by_id(lote_id)
    if not lote or lote.get("estado") != "Completado":
        raise HTTPException(status_code=404, detail="Plantilla no disponible")
    
    # For MVP, generate a simple Excel file
    # In production, this would create the actual Comiagro template
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Comiagro"
    
    # Add headers
    headers = ["NOMBRE USUARIO", "NIT USUARIO", "CIUDAD DE ENTREGA", "FACT NRO", 
               "FECHA", "FORMA DE PAGO", "PRODUCTO", "PRESENTACION", "CANTIDAD", 
               "VALOR UNITARIO", "TOTAL", "% IVA PRODUCTO"]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add sample data
    ws.cell(row=2, column=1, value="Empresa Cliente")
    ws.cell(row=2, column=2, value="901234567")
    ws.cell(row=2, column=3, value="Bogotá")
    ws.cell(row=2, column=4, value="F-001")
    ws.cell(row=2, column=5, value=datetime.now().strftime("%Y-%m-%d"))
    
    # Save to temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        temp_path = tmp.name
    
    return FileResponse(
        temp_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"plantilla_comiagro_lote_{lote_id}.xlsx"
    )

@app.get("/api/terceros/excepciones")
async def obtener_excepciones(
    estado: Optional[str] = None,
    lote_id: Optional[str] = None
):
    filters = {}
    if estado:
        filters['estado'] = estado
    if lote_id:
        filters['lote_id'] = lote_id
    
    excepciones = firebase_service.get_excepciones_dian(filters)
    
    # Convert timestamps
    for exc in excepciones:
        if 'fechaDeteccion' in exc:
            exc['fechaDeteccion'] = exc['fechaDeteccion'].isoformat()
        if 'fechaUltimaActualizacion' in exc:
            exc['fechaUltimaActualizacion'] = exc['fechaUltimaActualizacion'].isoformat()
    
    # Calculate statistics
    total = len(excepciones)
    pendientes = len([e for e in excepciones if e.get('estadoGestion') == 'Pendiente'])
    resueltas = total - pendientes
    
    return {
        "success": True,
        "data": {
            "excepciones": excepciones,
            "estadisticas": {
                "total": total,
                "pendientes": pendientes,
                "resueltas": resueltas
            }
        }
    }

@app.post("/api/terceros/excepciones/{excepcion_id}/{action}")
async def actualizar_estado_tercero(
    excepcion_id: str,
    action: str,
    payload: Optional[dict] = None
):
    # Get excepcion from Firestore
    excepciones = firebase_service.get_excepciones_dian()
    excepcion = next((e for e in excepciones if e.get('id') == excepcion_id), None)
    
    if not excepcion:
        raise HTTPException(status_code=404, detail="Excepción no encontrada")
    
    estado_anterior = excepcion.get('estadoValidacion')
    
    # Mapear acciones a estados
    accion_map = {
        "corregir": "Corregida",
        "crear": "En_Creacion_Manual",
        "ignorar": "Ignorada",
        "reintentar": "Reintentando"
    }
    
    if action not in accion_map:
        raise HTTPException(status_code=400, detail="Acción no válida")
    
    updates = {
        "estadoGestion": accion_map[action]
    }
    
    if payload:
        if 'notas' in payload:
            updates['notas'] = payload['notas']
        if 'datosCorreccion' in payload:
            updates['datosCorreccion'] = payload['datosCorreccion']
    
    # Update in Firestore
    success = firebase_service.update_excepcion_dian(excepcion_id, updates)
    
    if not success:
        raise HTTPException(status_code=500, detail="Error al actualizar excepción")
    
    return {
        "success": True,
        "message": f"Excepción actualizada: {action}",
        "data": {
            "excepcionId": excepcion_id,
            "accionAplicada": action,
            "estadoAnterior": estado_anterior,
            "estadoNuevo": accion_map[action],
            "fechaActualizacion": datetime.utcnow().isoformat()
        }
    }

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8080))
    uvicorn.run(app, host="0.0.0.0", port=port)
